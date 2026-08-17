import os
import json
import csv
import datetime
import http.server
import socketserver
from urllib.parse import parse_qs, urlparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# PORT is provided by the hosting platform (Render, Railway, etc). Falls back
# to 5000 for local development.
PORT = int(os.environ.get('PORT', 5000))

# DATA_DIR should point at a persistent disk mount when deployed, so the
# Excel/CSV files survive restarts and redeploys. Falls back to BASE_DIR for
# local development. On Render, set this to your mounted disk path (e.g.
# /var/data) in the service's environment variables.
DATA_DIR = os.environ.get('DATA_DIR', BASE_DIR)
os.makedirs(DATA_DIR, exist_ok=True)

EXCEL_FILE = os.path.join(DATA_DIR, 'contact_submissions.xlsx')
CSV_FILE = os.path.join(DATA_DIR, 'contact_submissions.csv')

HEADERS = ["ID", "Full Name", "Email Address", "Mobile Number", "Device Model", "Message", "Submission Date"]

def init_excel():
    """Create formatted Excel file if it does not exist yet."""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Submissions"
        
        # Write headers
        ws.append(HEADERS)
        
        # Style headers
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
        
        # Set column widths
        col_widths = [8, 22, 28, 18, 22, 40, 22]
        for idx, width in enumerate(col_widths, start=1):
            col_letter = openpyxl.utils.get_column_letter(idx)
            ws.column_dimensions[col_letter].width = width
            
        wb.save(EXCEL_FILE)
        print(f"[EXCEL INIT] Created new Excel file at: {EXCEL_FILE}")
        
    # Also initialize CSV backup file
    if not os.path.exists(CSV_FILE):
        with open(CSV_FILE, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(HEADERS)

def add_submission_to_excel(name, email, phone, device, message):
    """Append a new submission row into the Excel workbook and CSV file."""
    init_excel()
    
    # Load Excel Workbook
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Submissions"]
    
    next_id = ws.max_row
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S IST")
    
    row_data = [next_id, name, email, phone, device, message, timestamp]
    ws.append(row_data)
    
    # Format new row
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    current_row = ws.max_row
    for col_idx in range(1, len(row_data) + 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.border = thin_border
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True)
            
    wb.save(EXCEL_FILE)
    
    # Append to CSV file as well
    with open(CSV_FILE, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(row_data)
        
    print(f"[EXCEL ROW ADDED] Saved submission #{next_id} from '{name}' into {EXCEL_FILE}")
    return next_id, timestamp

def read_excel_submissions():
    """Read all rows from the Excel workbook as a list of dicts."""
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Submissions"]
    
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            records.append({
                'id': row[0],
                'name': str(row[1] or ''),
                'email': str(row[2] or ''),
                'phone': str(row[3] or ''),
                'device': str(row[4] or ''),
                'message': str(row[5] or ''),
                'created_at': str(row[6] or '')
            })
    return records

class ExcelHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=BASE_DIR, **kwargs)

    def do_GET(self):
        parsed = urlparse(self.path)
        
        # API: Return list of entries from Excel file
        if parsed.path == '/api/submissions':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            submissions = read_excel_submissions()
            self.wfile.write(json.dumps(submissions).encode('utf-8'))
            return
            
        # API: Download Excel File directly
        if parsed.path == '/api/download-excel':
            init_excel()
            if os.path.exists(EXCEL_FILE):
                with open(EXCEL_FILE, 'rb') as f:
                    content = f.read()
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', 'attachment; filename="contact_submissions.xlsx"')
                self.send_header('Content-Length', str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return

        return super().do_GET()

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == '/api/contact':
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length).decode('utf-8')
            
            try:
                data = json.loads(post_data)
            except Exception:
                data = {k: v[0] for k, v in parse_qs(post_data).items()}
            
            name = data.get('name', '').strip()
            email = data.get('email', '').strip()
            phone = data.get('phone', '').strip()
            device = data.get('device', '').strip()
            message = data.get('message', '').strip()

            if not name or not email or not message:
                self.send_response(400)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'status': 'error', 'message': 'Missing required fields'}).encode('utf-8'))
                return

            row_id, timestamp = add_submission_to_excel(name, email, phone, device, message)

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({
                'status': 'success',
                'id': row_id,
                'excel_file': 'contact_submissions.xlsx',
                'message': f'Saved dynamically to Excel file (Row #{row_id})'
            }).encode('utf-8'))
            return

        self.send_response(404)
        self.end_headers()

if __name__ == '__main__':
    init_excel()
    print(f"==================================================")
    print(f"  R SQUARE MOBILES — Excel File Storage Server")
    print(f"  Running on port: {PORT}")
    print(f"  Excel File: {EXCEL_FILE}")
    print(f"==================================================")
    
    with socketserver.TCPServer(("0.0.0.0", PORT), ExcelHandler) as httpd:
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\nServer stopped.")
